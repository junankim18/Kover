from openpyxl import load_workbook
from . import models
from .models import People
from django.shortcuts import render, redirect

load_wb = load_workbook(
    "C:/Users/Hi/Desktop/work/people.xlsx", data_only=True)
load_ws = load_wb['Sheet']

# all_values = []
# for row in load_ws.rows:
#     row_value = []
#     for cell in row:
#         row_value.append(cell.value)
#     all_values.append(row_value)


def crawl(request):
    for i in range(2, 101):
        if load_ws.cell(i, 3).value and load_ws.cell(i, 4).value:
            people = People(
                people_name=load_ws.cell(i, 1).value,
                people_type=load_ws.cell(i, 2).value,
                people_birth=load_ws.cell(i, 3).value,
                people_img=load_ws.cell(i, 4).value
            )
            people.save()
        elif load_ws.cell(i, 3).value == False and load_ws.cell(i, 4).value:
            people = People(
                people_name=load_ws.cell(i, 1).value,
                people_type=load_ws.cell(i, 2).value,
                people_img=load_ws.cell(i, 4).value
            )
            people.save()
        elif load_ws.cell(i, 3).value and load_ws.cell(i, 4).value == False:
            people = People(
                people_name=load_ws.cell(i, 1).value,
                people_type=load_ws.cell(i, 2).value,
                people_birth=load_ws.cell(i, 3).value,
                people_img='https://icons-for-free.com/iconfiles/png/512/anonymous+app+contacts+open+line+profile+user+icon-1320183042822068474.png',
            )
            people.save()
        elif load_ws.cell(i, 3).value == False and load_ws.cell(i, 4).value == False:
            people = People(
                people_name=load_ws.cell(i, 1).value,
                people_type=load_ws.cell(i, 2).value,
                people_img='https://icons-for-free.com/iconfiles/png/512/anonymous+app+contacts+open+line+profile+user+icon-1320183042822068474.png',
            )
            people.save()
    return render(request, 'kover/main.html')
